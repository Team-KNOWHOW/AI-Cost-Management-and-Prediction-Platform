import os

import openpyxl
from django.core.files.storage import FileSystemStorage
from django.shortcuts import render, redirect
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook

from .models import *
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from datetime import datetime

from bs4 import BeautifulSoup
import urllib.request as req

import pymysql
from django.conf import settings

MYDB = getattr(settings, "DATABASES", None)
MYDB_NAME = MYDB["default"]["NAME"]
MYDB_USER = MYDB["default"]["USER"]
MYDB_PWD = MYDB["default"]["PASSWORD"]
MYDB_HOST = MYDB["default"]["HOST"]
dbCon = pymysql.connect(host=MYDB_HOST, user=MYDB_USER, passwd=MYDB_PWD, database=MYDB_NAME)

board_path = "board/"


# Basic views
def home(request):  # 홈 화면.
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('/')

    url = "http://finance.naver.com/marketindex/"
    res = req.urlopen(url)
    soup = BeautifulSoup(res, "html.parser")
    price = soup.select_one("div.head_info > span.value").string
    graph = soup.select_one("#exchangeList > li.on > a.graph_img > img")['src']
    context['graph'] =graph
    context['price'] = price

    file_name = 'khmodel.h5'
    fs = FileSystemStorage(location='static/models')
    if (fs.exists(file_name)):
        file_name = 'khmodel.h5'
        context["model"] = file_name

    context["id"] = member_no
    context["user_id"] = member_id



    return render(request, 'home.html', context)


# User Register

def member_register(request):  # 회원가입 화면.
    return render(request, "registration/member_register.html")


@csrf_exempt
def member_id_check(request):  # 아이디 중복체크 기능.
    context = {}

    member_id = request.GET['user_id']
    rs = BUser.objects.filter(user_id=member_id).exists()

    if rs:
        context['flag'] = '1'
        context['result_msg'] = '이미 존재하는 아이디입니다.'
    else:
        context['flag'] = '0'
        context['result_msg'] = '사용가능한 아이디입니다.'

    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def member_insert(request):  # 회원등록 기능.
    context = {}

    member_id = request.GET['user_id']
    member_pwd = request.GET['psswd']
    member_name = request.GET['user_nm']
    member_phone_num = request.GET['phoneno']
    member_email = request.GET['email']

    rs = BUser.objects.create(user_id=member_id,
                              psswd=member_pwd,
                              user_nm=member_name,
                              email=member_email,
                              phoneno=member_phone_num,
                              usage_fg='1', )

    context['result_msg'] = '회원가입이 완료되었습니다.'

    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def member_login(request):  # 로그인 기능.
    context = {}

    member_id = request.GET['user_id']
    member_pwd = request.GET['psswd']

    if 'id' in request.session:
        context['flag'] = "1"
        context['result_msg'] = '이미 로그인 되어있는 아이디가 있습니다.'
    else:
        rs = BUser.objects.filter(user_id=member_id, psswd=member_pwd).exists()

        if rs:
            member = BUser.objects.get(user_id=member_id, psswd=member_pwd)
            member_no = member.id
            member.save()

            request.session['id'] = member_no
            request.session['user_id'] = member_id

            context['flag'] = "0"
            context['result_msg'] = '로그인이 완료되었습니다.'
        else:
            context['flag'] = "1"
            context['result_msg'] = '아이디 혹은 비밀번호가 일치하지 않습니다.'

    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def member_logout(request):  # 로그아웃 기능.
    context = {}

    request.session.flush()

    return redirect('main')


def member_check(request):  # 비밀번호 확인 화면.
    context = {}
    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('/')

    context["id"] = member_no
    context["user_id"] = member_id

    return render(request, "registration/member_check.html", context)


@csrf_exempt
def member_pwd_check(request):  # 비밀번호 확인 기능.
    context = {}

    member_pwd = request.GET['psswd']

    if 'id' in request.session:
        rs = BUser.objects.filter(psswd=member_pwd).exists()

        if rs:
            context['flag'] = "0"
            context['result_msg'] = '비밀번호가 확인되었습니다.'
        else:
            context['flag'] = "1"
            context['result_msg'] = '비밀번호가 일치하지 않습니다.'
    else:
        context['flag'] = "1"
        context['result_msg'] = '로그인 페이지로 이동합니다.'

    return JsonResponse(context, content_type="application/json")


def member_edit(request):  # 회원정보 변경화면.
    context = {}

    if 'id' in request.session:
        member_no = request.session['id']
        member = BUser.objects.get(id=member_no)

        context['id'] = member.id
        context['user_id'] = member.user_id
        context['user_nm'] = member.user_nm
        context['phoneno'] = member.phoneno
        context['email'] = member.email

        context['flag'] = "0"
        context['result_msg'] = '회원정보 수정가능합니다.'

        return render(request, "registration/member_edit.html", context)

    else:
        return redirect('/')


@csrf_exempt
def member_update(request):  # 회원정보 변경 기능.
    context = {}

    member_req = request.GET
    member_id = member_req.get('user_id')  # url에 포함되어있지 않으면 None 반환.
    member_pwd = member_req.get('psswd')
    member_name = member_req.get('user_nm')
    member_phone_num = member_req.get('phoneno')
    member_email = member_req.get('email')

    member = BUser.objects.get(user_id=member_id)

    if member_pwd is not None:
        member.psswd = member_pwd

    if member_name is not None:
        member.user_nm = member_name

    if member_phone_num is not None:
        member.phoneno = member_phone_num

    if member_email is not None:
        member.email = member_email

    member.save()

    context['result_msg'] = '회원정보 변경이 완료되었습니다.'

    return JsonResponse(context, content_type="application/json")


# *********************************************************************************************************************
# 거래처 코드 시작
# *********************************************************************************************************************

def b_bizpartner(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    strsql = "SELECT a.*, b.* ,c.*, d.* " + \
             "FROM (SELECT *FROM  b_bizpartner WHERE usage_fg='Y') a " + \
             "LEFT JOIN b_co b ON a.co_id=b.id " + \
             "LEFT JOIN (SELECT id, code_cd, cd_nm FROM cb_code_dtl WHERE type_cd ='country' ) c ON a.unitcn_id=c.id " + \
             "LEFT JOIN (SELECT id, code_cd, cd_nm FROM cb_code_dtl WHERE type_cd='currency') d ON a.unitcur_id=d.id "
    rsBizpartner = BBizpartner.objects.raw(strsql)
    context["rsBizpartner"] = rsBizpartner

    rsCo = BCo.objects.filter(usage_fg='Y')
    rsUnitCur = CbCodeDtl.objects.filter(type_cd='currency', usage_fg='Y')
    rsUnitCn = CbCodeDtl.objects.filter(type_cd='country', usage_fg='Y')
    rsBizpartnerstat = BBizpartner.objects.filter(usage_fg='Y')

    context['rsBizpartnerstat'] = rsBizpartnerstat
    context["rsCo"] = rsCo
    context["rsUnitCur"] = rsUnitCur
    context["rsUnitCn"] = rsUnitCn

    context["rsHeader"] = rsBizpartner

    context["title"] = "거래처"
    context["result_msg"] = "거래처"

    return render(request, board_path + "b_bizpartner.html", context)


def b_co(request):  # 법인정보
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    strsql = "SELECT a.*, b.*, c.* " + \
             "FROM (SELECT *FROM  b_co WHERE usage_fg='Y') a " + \
             "LEFT JOIN (SELECT id, code_cd, cd_nm FROM cb_code_dtl WHERE type_cd ='country' ) b ON a.unitcn_id=b.id " + \
             "LEFT JOIN (SELECT id, code_cd, cd_nm FROM cb_code_dtl WHERE type_cd='currency') c ON a.unitcur_id=c.id "
    rsCo = BCo.objects.raw(strsql)
    context["rsCo"] = rsCo

    rsUnitCur = CbCodeDtl.objects.filter(type_cd='currency', usage_fg='Y')
    rsUnitCn = CbCodeDtl.objects.filter(type_cd='country', usage_fg='Y')
    context["rsUnitCur"] = rsUnitCur
    context["rsUnitCn"] = rsUnitCn
    context["title"] = "법인정보"
    context["result_msg"] = "법인정보"

    return render(request, board_path + "b_co.html", context)


def b_bizarea(request):  # 사업장
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    strsql = "SELECT a.*, b.* ,c.*, d.* " + \
             "FROM (SELECT *FROM  b_bizarea WHERE usage_fg='Y') a " + \
             "LEFT JOIN b_co b ON a.co_id=b.id " + \
             "LEFT JOIN (SELECT id, code_cd, cd_nm FROM cb_code_dtl WHERE type_cd ='country' ) c ON a.unitcn_id=c.id " + \
             "LEFT JOIN (SELECT id, code_cd, cd_nm FROM cb_code_dtl WHERE type_cd='currency') d ON a.unitcur_id=d.id "
    rsBizarea = BBizarea.objects.raw(strsql)
    context["rsBizarea"] = rsBizarea

    rsCo = BCo.objects.filter(usage_fg='Y')
    rsUnitCur = CbCodeDtl.objects.filter(type_cd='currency', usage_fg='Y')
    rsUnitCn = CbCodeDtl.objects.filter(type_cd='country', usage_fg='Y')
    context["rsCo"] = rsCo
    context["rsUnitCur"] = rsUnitCur
    context["rsUnitCn"] = rsUnitCn
    context["title"] = "사업장"
    context["result_msg"] = "사업장"
    return render(request, board_path + "b_bizarea.html", context)


def b_bizunit(request):  # 사업부
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    rsHeader = BBizunit.objects.filter(usage_fg='Y')
    rsuserid = BUser.objects.filter(usage_fg='Y')  # user_id때문에

    context["title"] = "사업부"
    context["result_msg"] = "사업부"
    context["rsHeader"] = rsHeader
    context["rsuserid"] = rsuserid  # user_id

    return render(request, board_path + "b_bizunit.html", context)


def b_factory(request):  # 공장
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    # print(typecd)

    rsHeader = BFactory.objects.filter(usage_fg='Y')

    context["rsHeader"] = rsHeader

    context["title"] = "공장"
    context["result_msg"] = "공장"

    return render(request, board_path + "b_factory.html", context)


def codemanage(request):  # 통합코드 관리
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    if 'type_cd' in request.GET:
        typecd = request.GET['type_cd']
        rsCode = CbCodeDtl.objects.filter(type_cd=typecd, usage_fg='Y')
    else:
        typecd = None
        rsCode = None

    context["type_cd"] = typecd

    rsHeader = CbCodeHdr.objects.filter(usage_fg='Y')
    context["rsHeader"] = rsHeader
    context["rsCode"] = rsCode

    return render(request, "board/codemanage.html", context)


# 상세정보 html 만들어야함 일단 보류
def code_view(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    codeid = request.GET['codeid']
    rsCode = CbCodeDtl.objects.get(id=codeid)

    print(rsCode)

    context["type_cd"] = rsCode.type_cd
    context["code_cd"] = rsCode.code_cd
    context["code_name"] = rsCode.cd_nm

    context["result_msg"] = "Code detail"
    return render(request, "board/codeview.html", context)


def b_item(request):  # 품목마스터
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    context['flag'] = '0'
    context['result_msg'] = '품목코드 관리'

    itemcode = ''
    if 'itemcode' in request.GET:
        itemcode = request.GET['itemcode']
    itemname = ''
    if 'itemname' in request.GET:
        itemname = request.GET['itemname']
    itemspec = ''
    if 'itemspec' in request.GET:
        itemspec = request.GET['itemspec']

    if itemcode != "" and itemname != "" and itemspec != "":
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y' " \
                 "AND item_cd LIKE '%%" + itemcode + "%%' AND item_nm LIKE '%%" + itemname + "%%' AND item_spec LIKE '%%" + itemspec + "%%') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "
        rsItem = BItem.objects.raw(strSql)

    elif itemcode != "" and itemname != "":
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y' " \
                 "AND item_cd LIKE '%%" + itemcode + "%%' AND item_nm LIKE '%%" + itemname + "%%') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "
        rsItem = BItem.objects.raw(strSql)

    elif itemcode != "" and itemspec != "":
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y' " \
                 "AND item_cd LIKE '%%" + itemcode + "%%' AND item_spec LIKE '%%" + itemspec + "%%') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "
        rsItem = BItem.objects.raw(strSql)

    elif itemname != "" and itemspec != "":
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y' " \
                 "AND item_nm LIKE '%%" + itemname + "%%' AND item_spec LIKE '%%" + itemspec + "%%') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "
        rsItem = BItem.objects.raw(strSql)

    elif itemcode != "":
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y' AND item_cd LIKE '%%" + itemcode + "%%') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "

        rsItem = BItem.objects.raw(strSql)

    elif itemname != "":
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y' AND item_nm LIKE '%%" + itemname + "%%') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "

        rsItem = BItem.objects.raw(strSql)

    elif itemspec != "":
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y' AND item_spec LIKE '%%" + itemspec + "%%') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "

        rsItem = BItem.objects.raw(strSql)

    # rsItem = BItem.objects.filter(usage_fg='Y')
    else:
        strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN (SELECT id, code_cd AS unit_cd, cd_nm AS unit_name FROM cb_code_dtl WHERE type_cd = 'unit') c  ON a.unit_id = c.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id " + \
                 "LEFT JOIN b_itemaccnt e ON a.itemaccnt_id = e.id "

        rsItem = BItem.objects.raw(strSql)

    rsFactory = BFactory.objects.filter(usage_fg='Y')
    rsUnit = CbCodeDtl.objects.filter(type_cd='unit', usage_fg='Y')
    rsItemgrp = BItemgrp.objects.filter(usage_fg='Y')
    rsItemaccnt = BItemaccnt.objects.filter(usage_fg='Y')

    context["rsItem"] = rsItem
    context["rsItemgrp"] = rsItemgrp
    context["rsItemaccnt"] = rsItemaccnt
    context["rsFactory"] = rsFactory
    context["rsUnit"] = rsUnit

    return render(request, 'board/b_item.html', context)


def b_itemaccnt(request):  # 품목계정
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    rsItemaccnt = BItemaccnt.objects.filter(usage_fg='Y')

    context["rsItemaccnt"] = rsItemaccnt

    context["flag"] = "0"
    context["result_msg"] = "품목계정"

    return render(request, 'board/b_itemaccnt.html', context)


def b_itemgrp(request):  # 품목그룹
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    rsItemgrp = BItemgrp.objects.filter(usage_fg='Y')

    context["rsItemgrp"] = rsItemgrp

    context["flag"] = "0"
    context["result_msg"] = "품목그룹"

    return render(request, 'board/b_itemgrp.html', context)



def b_bom(request): # BOM
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    # 컨택스트 변수 초기화해주고
    context['itemid'] = 0
    context['itemcd'] = ""
    context['itemname'] = ""
    context['itemspec'] = ""
    context['registerdate'] = ""
    itemid = "0"

    # request안에 품폭id 가 있다면
    if 'itemid' in request.GET:
        itemid = request.GET['itemid']
        context['itemid'] = itemid
        # 필터로 걸러서 rsTmp로 담아주고 rsTmp객체의 속성들을 context변수에 넣어준다.
        if BItem.objects.filter(id=itemid).exists():
            rsTmp = BItem.objects.get(id=itemid)
            context['itemcd'] = rsTmp.item_cd
            context['itemname'] = rsTmp.item_nm
            context['itemspec'] = rsTmp.item_spec
            if rsTmp.insrt_dt:
                context['registerdate'] = rsTmp.insrt_dt
            else:
                context['registerdate'] = "No data of RegisterDate"
        else:
            print("nothing ")

    # 두번째 항목들 초기화.
    bomid = "0"
    context['moitembase'] = 0.0
    context['jaitembase'] = 0.0
    context['unitproduct'] = '단위'
    context['lossproduct'] = 0.0
    context['demandamt'] = 0.0
    context['startdate'] = ''
    context['enddate'] = ''

    # 두번째, request안에 bomid인 객체를 찾아서
    if 'bomid' in request.GET:
        bomid = request.GET['bomid']
        if BBom.objects.filter(id=bomid).exists():
            rsTmp2 = BBom.objects.get(id=bomid)
            context['moitembase'] = rsTmp2.moitem_base
            context['jaitembase'] = rsTmp2.jaitem_base
            context['unitproduct'] = rsTmp2.unit_product
            context['lossproduct'] = rsTmp2.loss_product
            context['demandamt'] = rsTmp2.demand_amt
            context['startdate'] = rsTmp2.start_dt
            context['enddate'] = rsTmp2.end_dt
        else:
            print("여기서 에러가 계속 나타나는 중")

    # 품폭그룹 아이디, 품목코드, 품목규격가져와서 저장.
    itemgrpid = ""
    if 'itemgrpid' in request.GET:
        itemgrpid = request.GET['itemgrpid']

    searchcode = ""
    if 'itemcode' in request.GET:
        searchcode = request.GET['itemcode']

    searchspec = ""
    if 'itemspec' in request.GET:
        searchspec = request.GET['itemspec']

    # like문 Q
    if searchcode != "":
        rsItem = BItem.objects.filter(Q(item_cd__contains=searchcode))[:100]
    elif searchspec != "":
        rsItem = BItem.objects.filter(Q(item_spec__contains=searchspec))[:100]
    elif itemgrpid != "":
        rsItem = BItem.objects.filter(itemgrp_id=itemgrpid)[:100]
    else:
        strsql = "SELECT a.*, b.*, d.*" + \
                 "FROM (SELECT * FROM b_item WHERE usage_fg = 'Y') a " + \
                 "LEFT JOIN b_factory b ON a.factory_id = b.id " + \
                 "LEFT JOIN b_itemgrp d ON a.itemgrp_id = d.id "
        rsItem = BItem.objects.raw(strsql)[:100]
    context['rsItem'] = rsItem

    rsBOM = BBom.objects.filter(top_id=itemid).select_related("item")

    context["rsBOM"] = rsBOM

    rsItemgrp = BItemgrp.objects.filter()
    context["rsItemgrp"] = rsItemgrp

    context['bomid'] = bomid
    context["itemgrpid"] = itemgrpid
    context["title"] = "BOM"
    context["result_msg"] = "BOM "

    return render(request, board_path + "b_bom.html", context)


@csrf_exempt
def bom_create(request):
    context = {}

    itemid = request.GET['itemid']

    if BBom.objects.filter(item_id=itemid, parent_id=0).exists():
        # print("already existed")
        context["flag"] = "1"
        context["result_msg"] = "이미 BOM TREE상에 존재합니다."
        return JsonResponse(context, content_type="application/json")
    else:
        BBom.objects.create(bom_type='MBOM',
                            item_id=itemid,
                            parent_id=0,
                            top_id=itemid,
                            bom_order=1,
                            bom_level=0,
                            leaf_fg='0',
                            moitem_base=0.0,
                            jaitem_base=0.0,
                            unit_product='',
                            demand_amt=0.0,
                            free_fg='0',
                            loss_product=0.0,
                            start_dt='',
                            end_dt='',
                            register_dt=datetime.now(),
                            usage_fg='Y')
        # rsItem에 아이디에 해당하는 품목을 저장하고 bomflag를 1로 바꿔주고 저장
        rsItem = BItem.objects.get(id=itemid)
        rsItem.bom_fg = '1'
        rsItem.save()

        context["flag"] = "0"
        context["result_msg"] = "Top level 등록 성공..."
        return JsonResponse(context, content_type="application/json")


@csrf_exempt
def bomitem_read(request):
    # 품목 조회를 클릭 시
    context = {}

    bomid = request.GET['bomid']
    itmtext = request.GET['itmtext']

    if itmtext == "":
        rsItem = BItem.objects.filter(usage_fg='Y')[:10]
    else:
        rsItem = BItem.objects.filter(Q(item_cd__contains=itmtext) | Q(item_spec__contains=itmtext))[:10]

    itmstr = ""
    # 품목아이템이 잘 불러와졌다면,
    if rsItem:
        for i in rsItem:
            # f를 붙여서 스크립트 명령어 표시
            itmstr += f"<div><i class='icofont-plus-square' style='margin-right:20px;' itemid='{i.id}' bomid='{bomid}' flag='add' onclick='pickBOMItem(this)'></i>  " + \
                      f"<i class='icofont-check' style='margin-right:20px;' itemid='{i.id}' bomid='{bomid}' flag='update' onclick='pickBOMItem(this)'></i> " + \
                      f"<span>{i.item_cd} - {i.item_spec} </span></div>"
    else:
        itmstr = "<div>No item searched...</div>"

    context["itmstr"] = itmstr
    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def bomitem_pick(request):
    context = {}

    bomid = request.GET['bomid']
    itemid = request.GET['itemid']
    flag = request.GET['flag']

    if flag == 'add':
        rsTmp = BBom.objects.get(id=bomid)
        bomorder = rsTmp.bom_order
        bomlevel = rsTmp.bom_level
        topid = rsTmp.top_id
        # 자신은 0이고 되고 밑은 1이되는 원리
        rsTmp.leaf_fg = '0'
        rsTmp.save()

        BBom.objects.create(bom_type='MBOM',
                            item_id=itemid,
                            parent_id=bomid,
                            top_id=topid,
                            # 정전개형 표현 하기 위해 bom_order
                            bom_order=bomorder + 1,
                            bom_level=bomlevel + 1,
                            leaf_fg='1',
                            moitem_base=0.0,
                            jaitem_base=0.0,
                            unit_product='',
                            demand_amt=0.0,
                            free_fg='0',
                            loss_product=0.0,
                            start_dt='',
                            end_dt='',
                            register_dt=datetime.now(),
                            usage_fg='Y')

        context["flag"] = "0"
        context["result_msg"] = "BOM child added..."
        return JsonResponse(context, content_type="application/json")

    elif flag == 'update':
        rsTmp = BBom.objects.get(id=bomid)
        rsTmp.item_id = itemid
        rsTmp.save()

        context["flag"] = "0"
        context["result_msg"] = "BOM item updated..."
        return JsonResponse(context, content_type="application/json")
    else:
        # 논리적으로 Error를 나타내는 부분 flag=1
        context["flag"] = "1"
        context["result_msg"] = "Nothing..."
        return JsonResponse(context, content_type="application/json")


@csrf_exempt
def bom_update(request):
    context = {}

    bomid = request.GET['bomid']
    flag = request.GET['flag']
    bvalue = request.GET['bvalue']

    rsTmp = BBom.objects.get(id=bomid)
    if flag == 'mobase':
        rsTmp.moitem_base = bvalue
        rsTmp.save()
    elif flag == 'jabase':
        rsTmp.jaitem_base = bvalue
        rsTmp.save()
    elif flag == 'unit':
        rsTmp.unit_product = bvalue
        rsTmp.save()
    elif flag == 'loss':
        rsTmp.loss_product = bvalue
        rsTmp.save()
    elif flag == 'demand':
        rsTmp.demand_amt = bvalue
        rsTmp.save()
    elif flag == 'sdate':
        rsTmp.start_dt = bvalue
        rsTmp.save()
    elif flag == 'edate':
        rsTmp.end_dt = bvalue
        rsTmp.save()
    else:
        context["flag"] = "1"
        context["result_msg"] = "Nothing updated..."
        return JsonResponse(context, content_type="application/json")

    context["flag"] = "0"
    context["result_msg"] = "BOM updated..."
    return JsonResponse(context, content_type="application/json")


# *********************************************************************************************************************
# BOM 코드 끝
# *********************************************************************************************************************

# *********************************************************************************************************************
# 작업장 코드 시작
# *********************************************************************************************************************

def b_workcenter(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    wc_query = "SELECT a.*, b.* " + \
               "FROM (SELECT * FROM b_workcenter WHERE usage_fg='Y') a " + \
               "LEFT JOIN cb_cost_center b ON a.cstctr_id=b.id"

    rsWorkcenter = BWorkcenter.objects.raw(wc_query)
    rsCostcenter = CbCostCenter.objects.filter(usage_fg='Y')

    context["title"] = "작업장"
    context["result_msg"] = "작업장"
    context['rsWorkcenter'] = rsWorkcenter
    context['rsCostcenter'] = rsCostcenter

    return render(request, board_path + 'b_workcenter.html', context)


def cb_cost_center(request):  # 코스트센터
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    # print(typecd)

    cc_query = "SELECT a.*, b.*, c.*, d.* " + \
               "FROM (SELECT * FROM cb_cost_center WHERE usage_fg='Y') a " + \
               "LEFT JOIN b_bizarea b ON a.bizarea_id=b.id " + \
               "LEFT JOIN b_bizunit c ON a.bizunit_id=c.id " + \
               "LEFT JOIN b_factory d ON a.factory_id=d.id"

    rsCostcenter = CbCostCenter.objects.raw(cc_query)
    rsBizarea = BBizarea.objects.filter(usage_fg='Y')
    rsBizunit = BBizunit.objects.filter(usage_fg='Y')
    rsFactory = BFactory.objects.filter(usage_fg='Y')

    context["title"] = "코스트센터"
    context["result_msg"] = "코스트센터"
    context["rsCostcenter"] = rsCostcenter
    context["rsFactory"] = rsFactory
    context["rsBizarea"] = rsBizarea
    context["rsBizunit"] = rsBizunit

    return render(request, board_path + 'cb_cost_center.html', context)


def b_costeleaccnt(request):  # 원가요소계정
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    rsCosteleaccnt = BCosteleaccnt.objects.filter(usage_fg='Y')

    context["rsCosteleaccnt"] = rsCosteleaccnt

    return render(request, 'board/b_costeleaccnt.html', context)


# *********************************************************************************************************************
# 2단계 제조비용 코드 시작
# *********************************************************************************************************************

def cc_manucost_if(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    strSql = "SELECT  a.*, b.*, c.*, d.*, e.* " \
             "FROM (SELECT * FROM cc_manucost_if) a " \
             "LEFT JOIN b_co b ON a.co_cd = b.co_cd " \
             "LEFT JOIN cb_cost_center c ON a.cstctr_cd = c.cstctr_cd " \
             "LEFT JOIN b_costeleaccnt d ON a.costeleaccnt_cd = d.accnt_cd " \
             "LEFT JOIN b_version e ON a.version_cd = e.version_cd"
    rsManucost = CcManucostIf.objects.raw(strSql)
    context["rsManucost"] = rsManucost

    rsCo = BCo.objects.filter(usage_fg='Y')
    rsCstctr = CbCostCenter.objects.filter(usage_fg='Y')
    rsCosteleaccnt = BCosteleaccnt.objects.filter(usage_fg='Y')
    rsVersion = BVersion.objects.all()

    context["rsCo"] = rsCo
    context["rsCstctr"] = rsCstctr
    context["rsCosteleaccnt"] = rsCosteleaccnt
    context["rsVersion"] = rsVersion

    return render(request, 'board2/cc_manucost_if.html', context)


@csrf_exempt
def manucosttemplate_download(request):
    context = {}

    strsql1 = "SHOW TABLES LIKE 'cc_manucost_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW FULL COLUMNS FROM cc_manucost_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

        idx = 1

        bookin = Workbook()
        sheet_in = bookin.active

        for i in rsColumns:
            sheet_in.cell(row=1, column=idx).value = i[1]
            sheet_in.cell(row=2, column=idx).value = i[8]
            idx += 1

        filename = "static/datatemplates/manucost.xlsx"
        bookin.save(filename)
        bookin.close()

        context["template_url"] = "/static/datatemplates/manucost.xlsx"

    else:
        context["flag"] = "1"
        context["result_msg"] = "품목코드 테이블 없음... "
        return JsonResponse(context, content_type="application/json")

    context["flag"] = "0"
    context["result_msg"] = "Template downloaded... "

    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def manucostdata_upload(request):
    context = {}

    if request.method == "POST":

        try:
            uploaded_file = request.FILES['ufile']
            name_old = uploaded_file.name
            # name_ext = os.path.splitext(name_old)[1]
            # name_new = "cc_manucost"
        except:
            print("파일선택x")
            return redirect('board:cc_manucost_if')
        file_name = name_old
        # file_name = name_new + name_ext

        fs = FileSystemStorage(location='static/dataupload/cc_manucost')
        # if (fs.exists(file_name)):
        #     fs.delete(file_name)
        name = fs.save(file_name, uploaded_file)

        context['upload_url'] = fs.url(name)
        context['upload_flag'] = 'USuccess'

    else:
        return redirect('board:cc_manucost_if')

    strsql1 = "SHOW TABLES LIKE 'cc_manucost_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW COLUMNS FROM cc_manucost_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

    else:
        print("테이블 없음")
        return redirect('board:cc_manucost_if')

    max_col = len(rsColumns)

    timenow = datetime.now()

    filename = 'static/dataupload/cc_manucost/' + file_name

    if not os.path.isfile(filename):
        print("저장안됨")
        return redirect('board:cc_manucost_if')
    else:
        book = openpyxl.load_workbook(filename, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

    if max_row > 2:
        try:
            for j in range(3, max_row + 1):
                lstTmp = []
                strbottom = ""

                for i in range(1, max_col + 1):
                    valTmp = sheet.cell(row=j, column=i).value

                    lstTmp.append(valTmp)

                    if valTmp == '':
                        strbottom += "default,"
                    elif valTmp == None:
                        strbottom += "default,"
                    else:
                        strbottom += "'" + str(valTmp) + "',"

                strbottom = strbottom[:-1]

                strSql = "INSERT INTO cc_manucost_if VALUES (" + strbottom + ")"

                cursor = dbCon.cursor()
                cursor.execute(strSql)
                rows = cursor.fetchone()
                cursor.close()

            dbCon.commit()
            book.close()
        except:
            print("엑셀오류")
            return redirect('board:cc_manucost_if')
    else:
        print("데이터x")
        return redirect('board:cc_manucost_if')

    return redirect('board:cc_manucost_if')


# *********************************************************************************************************************
# 제조비용 코드 끝
# *********************************************************************************************************************


# *********************************************************************************************************************
# 재료비 코드 시작
# *********************************************************************************************************************

def cc_materialcost_if(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    strSql = "SELECT  a.*, b.*, c.*, d.*, e.*, f.*, h.*, i.*, j.*" \
             "FROM (SELECT * FROM cc_materialcost_if) a " \
             "LEFT JOIN (SELECT factory_cd f_cd, factory_nm FROM b_factory) b ON a.factory_cd = b.f_cd " \
             "LEFT JOIN b_co c ON a.co_cd = c.co_cd " \
             "LEFT JOIN b_workcenter d ON a.workcenter_cd = d.workcenter_cd " \
             "LEFT JOIN b_costeleaccnt e ON a.costeleaccnt_cd = e.accnt_cd " \
             "LEFT JOIN (SELECT b1.item_cd moitem, b2.item_cd jaitem FROM b_bom b1, b_bom b2 " \
             "WHERE b1.item_id = b2.parent_id) f ON f.moitem = a.bom_cd " \
             "LEFT JOIN (SELECT item_nm janame, item_cd FROM b_item) j on f.jaitem = j.item_cd " \
             "LEFT JOIN b_version h ON a.version_cd = h.version_cd " \
             "LEFT JOIN b_item i ON a.bom_cd = i.item_cd"

    rsMaterialcost = CcMaterialcostIf.objects.raw(strSql)
    context["rsMaterialcost"] = rsMaterialcost

    rsCo = BCo.objects.filter(usage_fg='Y')
    rsFactory = BFactory.objects.filter(usage_fg='Y')
    rsWrkctr = BWorkcenter.objects.filter(usage_fg='Y')
    rsCosteleaccnt = BCosteleaccnt.objects.filter(usage_fg='Y')
    rsVersion = BVersion.objects.all()
    rsBom = BBom.objects.filter(usage_fg='Y', parent_id=0)

    context["rsCo"] = rsCo
    context["rsFactory"] = rsFactory
    context["rsWrkctr"] = rsWrkctr
    context["rsCosteleaccnt"] = rsCosteleaccnt
    context["rsBom"] = rsBom
    context["rsVersion"] = rsVersion


    return render(request, 'board2/cc_materialcost_if.html', context)


@csrf_exempt
def materialcosttemplate_download(request):
    context = {}

    strsql1 = "SHOW TABLES LIKE 'cc_materialcost_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW FULL COLUMNS FROM cc_materialcost_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

        idx = 1

        bookin = Workbook()
        sheet_in = bookin.active

        for i in rsColumns:
            sheet_in.cell(row=1, column=idx).value = i[1]
            sheet_in.cell(row=2, column=idx).value = i[8]
            idx += 1

        filename = "static/datatemplates/materialcost.xlsx"
        bookin.save(filename)
        bookin.close()

        context["template_url"] = "/static/datatemplates/materialcost.xlsx"

    else:
        context["flag"] = "1"
        context["result_msg"] = "품목코드 테이블 없음... "
        return JsonResponse(context, content_type="application/json")

    context["flag"] = "0"
    context["result_msg"] = "Template downloaded... "

    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def materialcostdata_upload(request):
    context = {}

    if request.method == "POST":

        try:
            uploaded_file = request.FILES['ufile']
            name_old = uploaded_file.name
            # name_ext = os.path.splitext(name_old)[1]
            # name_new = "cc_manucost"
        except:
            print("파일선택x")
            return redirect('board:cc_materialcost_if')
        file_name = name_old
        # file_name = name_new + name_ext

        fs = FileSystemStorage(location='static/dataupload/cc_materialcost')
        # if (fs.exists(file_name)):
        #     fs.delete(file_name)
        name = fs.save(file_name, uploaded_file)

        context['upload_url'] = fs.url(name)
        context['upload_flag'] = 'USuccess'

    else:
        return redirect('board:cc_materialcost_if')

    strsql1 = "SHOW TABLES LIKE 'cc_materialcost_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW COLUMNS FROM cc_materialcost_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

    else:
        print("테이블 없음")
        return redirect('board:cc_materialcost_if')

    max_col = len(rsColumns)

    timenow = datetime.now()

    filename = 'static/dataupload/cc_materialcost/' + file_name

    if not os.path.isfile(filename):
        print("저장안됨")
        return redirect('board:cc_materialcost_if')
    else:
        book = openpyxl.load_workbook(filename, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

    if max_row > 2:
        try:
            for j in range(3, max_row + 1):
                lstTmp = []
                strbottom = ""

                for i in range(1, max_col + 1):
                    valTmp = sheet.cell(row=j, column=i).value

                    lstTmp.append(valTmp)

                    if valTmp == '':
                        strbottom += "default,"
                    elif valTmp == None:
                        strbottom += "default,"
                    else:
                        strbottom += "'" + str(valTmp) + "',"

                strbottom = strbottom[:-1]

                strSql = "INSERT INTO cc_materialcost_if VALUES (" + strbottom + ")"

                cursor = dbCon.cursor()
                cursor.execute(strSql)
                rows = cursor.fetchone()
                cursor.close()

            dbCon.commit()
            book.close()
        except:
            print("엑셀오류")
            return redirect('board:cc_materialcost_if')
    else:
        print("데이터x")
        return redirect('board:cc_materialcost_if')

    return redirect('board:cc_materialcost_if')


# *********************************************************************************************************************
# 재료비 코드 끝
# *********************************************************************************************************************

# *********************************************************************************************************************
# 품목별 제조비용 코드 시작
# *********************************************************************************************************************

def cc_itempermanucost_if(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    strSql = "SELECT  a.*, b.*, c.*, d.*, e.*, f.* " \
             "FROM (SELECT * FROM cc_itempermanucost_if) a " \
             "LEFT JOIN b_co b ON a.co_cd = b.co_cd " \
             "LEFT JOIN b_costeleaccnt c ON a.costeleaccnt_cd = c.accnt_cd " \
             "LEFT JOIN b_bom d ON a.bom_cd = d.item_cd " \
             "LEFT JOIN b_version e ON a.version_cd = e.version_cd " \
             "LEFT JOIN b_item f ON a.bom_cd = f.item_cd "

    rsItempermanucost = CcItempermanucostIf.objects.raw(strSql)
    context["rsItempermanucost"] = rsItempermanucost

    rsCo = BCo.objects.filter(usage_fg='Y')
    rsCosteleaccnt = BCosteleaccnt.objects.filter(usage_fg='Y')
    rsBom = BBom.objects.filter(usage_fg='Y', parent_id=0)
    rsVersion = BVersion.objects.all()

    context["rsCo"] = rsCo
    context["rsCosteleaccnt"] = rsCosteleaccnt
    context["rsBom"] = rsBom
    context["rsVersion"] = rsVersion

    return render(request, 'board2/cc_itempermanucost_if.html', context)


@csrf_exempt
def itempermanucosttemplate_download(request):
    context = {}

    strsql1 = "SHOW TABLES LIKE 'cc_itempermanucost_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW FULL COLUMNS FROM cc_itempermanucost_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

        idx = 1

        bookin = Workbook()
        sheet_in = bookin.active

        for i in rsColumns:
            sheet_in.cell(row=1, column=idx).value = i[1]
            sheet_in.cell(row=2, column=idx).value = i[8]
            idx += 1

        filename = "static/datatemplates/itempermanucost.xlsx"
        bookin.save(filename)
        bookin.close()

        context["template_url"] = "/static/datatemplates/itempermanucost.xlsx"

    else:
        context["flag"] = "1"
        context["result_msg"] = "품목코드 테이블 없음... "
        return JsonResponse(context, content_type="application/json")

    context["flag"] = "0"
    context["result_msg"] = "Template downloaded... "

    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def itempermanucostdata_upload(request):
    context = {}

    if request.method == "POST":

        try:
            uploaded_file = request.FILES['ufile']
            name_old = uploaded_file.name
            # name_ext = os.path.splitext(name_old)[1]
            # name_new = "cc_manucost"
        except:
            print("파일선택x")
            return redirect('board:cc_itempermanucost_if')
        file_name = name_old
        # file_name = name_new + name_ext

        fs = FileSystemStorage(location='static/dataupload/cc_itempermanucost')
        # if (fs.exists(file_name)):
        #     fs.delete(file_name)
        name = fs.save(file_name, uploaded_file)

        context['upload_url'] = fs.url(name)
        context['upload_flag'] = 'USuccess'

    else:
        return redirect('board:cc_itempermanucost_if')

    strsql1 = "SHOW TABLES LIKE 'cc_itempermanucost_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW COLUMNS FROM cc_itempermanucost_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

    else:
        print("테이블 없음")
        return redirect('board:cc_itempermanucost_if')

    max_col = len(rsColumns)

    # timenow = datetime.now()

    filename = 'static/dataupload/cc_itempermanucost/' + file_name

    if not os.path.isfile(filename):
        print("저장안됨")
        return redirect('board:cc_itempermanucost_if')
    else:
        book = openpyxl.load_workbook(filename, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

    if max_row > 2:

        try:
            for j in range(3, max_row + 1):
                lstTmp = []
                strbottom = ""

                for i in range(1, max_col + 1):
                    valTmp = sheet.cell(row=j, column=i).value

                    lstTmp.append(valTmp)

                    if valTmp == '':
                        strbottom += "default,"
                    elif valTmp == None:
                        strbottom += "default,"
                    else:
                        strbottom += "'" + str(valTmp) + "',"

                strbottom = strbottom[:-1]

                strSql = "INSERT INTO cc_itempermanucost_if VALUES (" + strbottom + ")"

                cursor = dbCon.cursor()
                cursor.execute(strSql)
                rows = cursor.fetchone()
                cursor.close()

            dbCon.commit()
            book.close()

        except:
            print("엑셀오류")
            return redirect('board:cc_itempermanucost_if')
    else:
        print("데이터x")
        return redirect('board:cc_itempermanucost_if')

    return redirect('board:cc_itempermanucost_if')


# *********************************************************************************************************************
# 품목별 제조비용 코드 끝
# *********************************************************************************************************************

# *********************************************************************************************************************
# 제품원가수불 코드 시작
# *********************************************************************************************************************

def cc_productcostpayment_if(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    strSql = "SELECT  a.*, b.*, c.*, d.*, e.*, f.* " \
             "FROM (SELECT * FROM cc_productcostpayment_if) a " \
             "LEFT JOIN (SELECT factory_cd f_cd, factory_nm FROM b_factory) b ON a.factory_cd = b.f_cd " \
             "LEFT JOIN b_bom c ON a.bom_cd = c.item_cd " \
             "LEFT JOIN b_costeleaccnt d ON a.costeleaccnt_cd = d.accnt_cd " \
             "LEFT JOIN b_version e ON a.version_cd = e.version_cd " \
             "LEFT JOIN b_item f ON a.bom_cd = f.item_cd "

    rsProductcostpayment = CcProductcostpaymentIf.objects.raw(strSql)
    context["rsProductcostpayment"] = rsProductcostpayment

    rsFactory = BFactory.objects.filter(usage_fg='Y')
    rsCosteleaccnt = BCosteleaccnt.objects.filter(usage_fg='Y')
    rsBom = BBom.objects.filter(usage_fg='Y', parent_id=0)
    rsVersion = BVersion.objects.all()


    context["rsFactory"] = rsFactory
    context["rsCosteleaccnt"] = rsCosteleaccnt
    context["rsBom"] = rsBom
    context["rsVersion"] = rsVersion

    return render(request, 'board2/cc_productcostpayment_if.html', context)


@csrf_exempt
def productcostpaymenttemplate_download(request):
    context = {}

    strsql1 = "SHOW TABLES LIKE 'cc_productcostpayment_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW FULL COLUMNS FROM cc_productcostpayment_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

        idx = 1

        bookin = Workbook()
        sheet_in = bookin.active

        for i in rsColumns:
            sheet_in.cell(row=1, column=idx).value = i[1]
            sheet_in.cell(row=2, column=idx).value = i[8]
            idx += 1

        filename = "static/datatemplates/productcostpayment.xlsx"
        bookin.save(filename)
        bookin.close()

        context["template_url"] = "/static/datatemplates/productcostpayment.xlsx"

    else:
        context["flag"] = "1"
        context["result_msg"] = "품목코드 테이블 없음... "
        return JsonResponse(context, content_type="application/json")

    context["flag"] = "0"
    context["result_msg"] = "Template downloaded... "

    return JsonResponse(context, content_type="application/json")


@csrf_exempt
def productcostpaymentdata_upload(request):
    context = {}

    if request.method == "POST":

        try:
            uploaded_file = request.FILES['ufile']
            name_old = uploaded_file.name
            # name_ext = os.path.splitext(name_old)[1]
            # name_new = "cc_manucost"
        except:
            print("파일선택x")
            return redirect('board:cc_productcostpayment_if')
        file_name = name_old
        # file_name = name_new + name_ext

        fs = FileSystemStorage(location='static/dataupload/cc_productcostpayment')
        # if (fs.exists(file_name)):
        #     fs.delete(file_name)
        name = fs.save(file_name, uploaded_file)

        context['upload_url'] = fs.url(name)
        context['upload_flag'] = 'USuccess'

    else:
        return redirect('board:cc_productcostpayment_if')

    strsql1 = "SHOW TABLES LIKE 'cc_productcostpayment_if'"

    cursor1 = dbCon.cursor()
    cursor1.execute(strsql1)
    rsTmp = cursor1.fetchone()
    cursor1.close()

    rsColumns = None
    if rsTmp:
        strsql1 = "SHOW COLUMNS FROM cc_productcostpayment_if"

        cursor2 = dbCon.cursor()
        cursor2.execute(strsql1)
        rsColumns = cursor2.fetchall()
        cursor2.close()

    else:
        print("테이블 없음")
        return redirect('board:cc_productcostpayment_if')

    max_col = len(rsColumns)

    # timenow = datetime.now()

    filename = 'static/dataupload/cc_productcostpayment/' + file_name

    if not os.path.isfile(filename):
        print("저장안됨")
        return redirect('board:cc_productcostpayment_if')
    else:
        book = openpyxl.load_workbook(filename, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

    if max_row > 2:
        try:
            for j in range(3, max_row + 1):
                lstTmp = []
                strbottom = ""

                for i in range(1, max_col + 1):
                    valTmp = sheet.cell(row=j, column=i).value

                    lstTmp.append(valTmp)

                    if valTmp == '':
                        strbottom += "default,"
                    elif valTmp == None:
                        strbottom += "default,"
                    else:
                        strbottom += "'" + str(valTmp) + "',"

                strbottom = strbottom[:-1]

                strSql = "INSERT INTO cc_productcostpayment_if VALUES (" + strbottom + ")"

                cursor = dbCon.cursor()
                cursor.execute(strSql)
                rows = cursor.fetchone()
                cursor.close()

            dbCon.commit()
            book.close()
        except:
            print("엑셀오류")
            return redirect('board:cc_productcostpayment_if')
    else:
        print("데이터x")
        return redirect('board:cc_productcostpayment_if')

    return redirect('board:cc_productcostpayment_if')


# *********************************************************************************************************************
# 제품원가수불 코드 끝
# *********************************************************************************************************************


# *********************************************************************************************************************
# costbill 코드 시작
# *********************************************************************************************************************

def cc_costbill_if(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    return render(request, 'board2/cc_costbill_if.html', context)


# *********************************************************************************************************************
# costbill 코드 끝
# *********************************************************************************************************************

def cc_costbill1_if(request): #분석용
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id

    versioncd = request.GET.get('versioncd', '')
    if versioncd == '':
        versioncd = CcCostbill.objects.all().values_list('version_cd', flat=True).order_by('version_cd')[0]

    periodcd = request.GET.get('periodcd', '')
    if(periodcd == ''):
        periodcd = CcCostbill.objects.all().values_list('periodym_cd', flat=True).order_by('periodym_cd').last()

    itemcd = request.GET.get('itemcd', '')



    sql = "SELECT id, version_cd, periodym_cd,item_cd AS 모델명, proq*proamt_unit AS 매출, " \
          "bi_brm AS 기초제공품합계, proq AS 생산량, proamt_unit AS 단가, proq*proamt_unit AS 매출, " \
          "ic_dlvc+ic_dlfc+ic_idlc AS 당기투입노무비합, " \
          "ic_ohdvc+ic_ohdfd+ic_ohdfe+ic_idohc AS 당기투입제조경비합, " \
          "ic_ohdvc+ic_ohdfd+ic_ohdfe+ic_idohc+ ic_dlvc+ic_dlfc+ic_idlc+ic_arm  AS 당기투입합계, " \
          "ic_arm AS 당기투입재료비, " \
          "uc_dlc+ uc_idlc AS 당기사용노무비합, " \
          "uc_dohc+uc_idohc AS 당기사용제조경비합, " \
          "ra_rm AS 타계정으로대체재료비, " \
          "uc_srw AS 당기사용재료비, " \
          "uc_dlc+ uc_idlc+uc_dohc+uc_idohc +uc_srw AS 당기사용제조원가총액, " \
          "ei_erm+ei_elc+ei_eoh AS 기말제공품합계 " \
          "FROM cc_costbill WHERE version_cd= '" + versioncd + "' AND periodym_cd = "+ str(periodcd)

    sql1 = '''SELECT id, item_cd AS 모델명
            FROM cc_costbill group by item_cd;'''
    sql2 = '''SELECT id, version_cd
                FROM cc_costbill group by version_cd;'''

    if(itemcd != ''):
        sql = "SELECT id, version_cd, periodym_cd,item_cd AS 모델명, proq*proamt_unit AS 매출, " \
          "bi_brm AS 기초제공품합계, proq AS 생산량, proamt_unit AS 단가, proq*proamt_unit AS 매출, " \
          "ic_dlvc+ic_dlfc+ic_idlc AS 당기투입노무비합, " \
          "ic_ohdvc+ic_ohdfd+ic_ohdfe+ic_idohc AS 당기투입제조경비합, " \
          "ic_ohdvc+ic_ohdfd+ic_ohdfe+ic_idohc+ ic_dlvc+ic_dlfc+ic_idlc+ic_arm  AS 당기투입합계, " \
          "ic_arm AS 당기투입재료비, " \
          "uc_dlc+ uc_idlc AS 당기사용노무비합, " \
          "uc_dohc+uc_idohc AS 당기사용제조경비합, " \
          "ra_rm AS 타계정으로대체재료비, " \
          "uc_srw AS 당기사용재료비, " \
          "uc_dlc+ uc_idlc+uc_dohc+uc_idohc +uc_srw AS 당기사용제조원가총액, " \
          "ei_erm+ei_elc+ei_eoh AS 기말제공품합계 " \
          "FROM cc_costbill WHERE version_cd= '" + versioncd + "' AND periodym_cd = " + str(periodcd) + " AND item_cd = '" + itemcd + "'"

    rsCosbtill1 = CcCostbill.objects.raw(sql)
    rsItemcd = CcCostbill.objects.raw(sql1)
    rsVersioncd = CcCostbill.objects.raw(sql2)

    context["versioncd"] = versioncd
    context["itemcd"] = itemcd
    context["periodcd"] = periodcd

    context["rsCostbill1"] = rsCosbtill1
    context["rsItemcd"] = rsItemcd
    context["rsVersioncd"] = rsVersioncd


    return render(request, 'board3/cc_costbill1_if.html', context)


# *********************************************************************************************************************
# 차트 코드 시작
# *********************************************************************************************************************

def chart_if(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    return render(request, 'board3/chart_if.html', context)


def chart1(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    return render(request, 'board3/chart1.html', context)

def chart2(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    return render(request, 'board3/chart2.html', context)

def chart3(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    context["id"] = member_no
    context["user_id"] = member_id


    return render(request, 'board3/chart3.html', context)

def simulate(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')



    return render(request, 'board3/simulate.html', context)

@csrf_exempt
def insertSimulate(request):
    context = {}

    if request.session.has_key('id'):  # 로그인 되어있는 상태인지 체크.
        member_no = request.session['id']
        member_id = request.session['user_id']
    else:
        member_no = None
        member_id = None

        return redirect('board:home')

    variable = request.GET['variable']
    fixed = request.GET['fixed']
    material = request.GET['material']

    context['success'] = True
    context['message'] = '선박 삭제되었습니다.'

    rs = CaPridiction.objects.create(variableperc_cost=variable,
                                     fixedperc_cost=fixed,
                                     materialperc_cost=material)

    context['result_msg'] = '등록.'


    return JsonResponse(context, content_type='application/json')