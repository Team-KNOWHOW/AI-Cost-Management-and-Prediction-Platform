import openpyxl
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import *
from .serializers import *
from rest_framework.parsers import JSONParser
from rest_framework.decorators import api_view
from rest_framework import exceptions
from openpyxl import Workbook
import pymysql
from django.conf import settings
from .jw import *

MYDB = getattr(settings, "DATABASES", None)
MYDB_NAME = MYDB["default"]["NAME"]
MYDB_USER = MYDB["default"]["USER"]
MYDB_PWD = MYDB["default"]["PASSWORD"]
MYDB_HOST = MYDB["default"]["HOST"]
dbCon = pymysql.connect(host=MYDB_HOST, user=MYDB_USER, passwd=MYDB_PWD, database=MYDB_NAME)


@api_view(['GET', 'POST'])
@csrf_exempt
def co_list(request):
    if request.method == 'GET':
        query_set = BCo.objects.all()
        serializer = BcoSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BCo.objects.filter(co_cd=data['co_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BcoSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def co_detail(request, pk):
    obj = BCo.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BcoSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BcoSerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def bizarea_list(request):
    if request.method == 'GET':
        query_set = BBizarea.objects.all()
        serializer = BBizareaSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BBizarea.objects.filter(bizarea_cd=data['bizarea_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BBizareaSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def bizarea_detail(request, pk):
    obj = BBizarea.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BBizareaSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BBizareaSerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def bizunit_list(request):
    if request.method == 'GET':
        query_set = BBizunit.objects.all()
        serializer = BBizunitSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BBizunit.objects.filter(bizunit_cd=data['bizunit_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BBizunitSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def bizunit_detail(request, pk):
    obj = BBizunit.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BBizunitSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BBizunitSerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def factory_list(request):
    if request.method == 'GET':
        query_set = BFactory.objects.all()
        serializer = BFactorySerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BFactory.objects.filter(factory_cd=data['factory_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BFactorySerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def factory_detail(request, pk):
    obj = BFactory.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BFactorySerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BFactorySerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def workcenter_list(request):
    if request.method == 'GET':
        query_set = BWorkcenter.objects.all()
        serializer = BWorkcenterSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BWorkcenter.objects.filter(workcenter_cd=data['workcenter_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BWorkcenterSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def workcenter_detail(request, pk):
    obj = BWorkcenter.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BWorkcenterSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BWorkcenterSerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def bizpartner_list(request):
    if request.method == 'GET':
        query_set = BBizpartner.objects.all()
        serializer = BBizpartnerSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BBizpartner.objects.filter(bizpartner_cd=data['bizpartner_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BBizpartnerSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def bizpartner_detail(request, pk):
    obj = BBizpartner.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BBizpartnerSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BBizpartnerSerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        # print(serializer.errors.values())
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def itemaccnt_list(request):
    if request.method == 'GET':
        query_set = BItemaccnt.objects.all()
        serializer = BItemaccntSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BItemaccnt.objects.filter(itemaccnt_cd=data['itemaccnt_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BItemaccntSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def itemaccnt_detail(request, pk):
    obj = BItemaccnt.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BItemaccntSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BItemaccntSerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def itemgrp_list(request):
    if request.method == 'GET':
        query_set = BItemgrp.objects.all()
        serializer = BItemgrpSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BItemgrp.objects.filter(itemgrp_cd=data['itemgrp_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("DuplicateCode")

        serializer = BItemgrpSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def itemgrp_detail(request, pk):
    obj = BItemgrp.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BItemgrpSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)
        serializer = BItemgrpSerializer(obj, data=data)
        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def code_hdr_list(request):
    if request.method == 'GET':
        query_set = CbCodeHdr.objects.all()
        serializer = CbCodeHdrSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if CbCodeHdr.objects.filter(type_cd=data['type_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Type Code")

        if CbCodeHdr.objects.filter(type_nm=data['type_nm'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Type Name")

        serializer = CbCodeHdrSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def code_hdr_detail(request, pk):
    obj = CbCodeHdr.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = CbCodeHdrSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)

        if CbCodeHdr.objects.filter(type_nm=data['type_nm'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Type Name")

        serializer = CbCodeHdrSerializer(obj, data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        if CbCodeDtl.objects.filter(type_cd=obj.type_cd, usage_fg='Y').exists():
            raise exceptions.ParseError("Can't delete root code")

        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def code_dtl_list(request):
    if request.method == 'GET':
        query_set = CbCodeDtl.objects.all()
        serializer = CbCodeDtlSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if CbCodeDtl.objects.filter(cd_nm=data['cd_nm'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Code Name")

        serializer = CbCodeDtlSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def code_dtl_detail(request, pk):
    obj = CbCodeDtl.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = CbCodeDtlSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)

        if CbCodeDtl.objects.filter(cd_nm=data['cd_nm'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Code Name")

        serializer = CbCodeDtlSerializer(obj, data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def cstctr_list(request):
    if request.method == 'GET':
        query_set = CbCostCenter.objects.all()
        serializer = CbCostCenterSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if CbCostCenter.objects.filter(cstctr_cd=data['cstctr_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Code")

        if CbCostCenter.objects.filter(cstctr_nm=data['cstctr_nm'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Name")

        serializer = CbCostCenterSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def cstctr_detail(request, pk):
    obj = CbCostCenter.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = CbCostCenterSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)

        serializer = CbCostCenterSerializer(obj, data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def item_list(request):
    if request.method == 'GET':
        query_set = BItem.objects.all()
        serializer = BItemSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        serializer = BItemSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def item_detail(request, pk):
    obj = BItem.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BItemSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)

        serializer = BItemSerializer(obj, data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def costeleaccnt_list(request):
    if request.method == 'GET':
        query_set = BCosteleaccnt.objects.all()
        serializer = BCosteleaccntSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        if BCosteleaccnt.objects.filter(pl_cd=data['pl_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate PL Code")

        if BCosteleaccnt.objects.filter(accnt_cd=data['accnt_cd'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Account Code")

        serializer = BCosteleaccntSerializer(data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def costeleaccnt_detail(request, pk):
    obj = BCosteleaccnt.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = BCosteleaccntSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)

        serializer = BCosteleaccntSerializer(obj, data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


@api_view(['GET', 'POST'])
@csrf_exempt
def cc_manucost_if(request):
    if request.method == 'GET':  # Excel Template Download

        strsql1 = "SHOW TABLES LIKE 'cc_manucost_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        rsColumns = None
        if rsTmp:
            strsql1 = "SHOW FULL COLUMNS FROM cc_manucost_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

            idx = 1

            bookin = Workbook()
            sheet_in = bookin.active

            for i in rsColumns:
                sheet_in.cell(row=1, column=idx).value = i[1]
                sheet_in.cell(row=2, column=idx).value = i[8]
                idx += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="manucost.xlsx"'

            bookin.save(response)
            bookin.close()

        return response

    elif request.method == 'POST':  # Excel Template Upload
        file = request.FILES['input_file']

        strsql1 = "SHOW TABLES LIKE 'cc_manucost_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        if rsTmp:
            strsql1 = "SHOW COLUMNS FROM cc_manucost_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

        max_col = len(rsColumns)
        book = openpyxl.load_workbook(file, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

        if max_row > 2:
            try:
                for j in range(3, max_row + 1):
                    lstTmp = []
                    strbottom = ""

                    for i in range(1, max_col + 1):
                        valTmp = sheet.cell(row=j, column=i).value

                        lstTmp.append(valTmp)

                        if valTmp == '':
                            strbottom += "default,"
                        elif valTmp == None:
                            strbottom += "default,"
                        else:
                            strbottom += "'" + str(valTmp) + "',"

                    strbottom = strbottom[:-1]

                    strSql = "INSERT INTO cc_manucost_if VALUES (" + strbottom + ")"

                    cursor = dbCon.cursor()
                    cursor.execute(strSql)
                    rows = cursor.fetchone()
                    cursor.close()

                dbCon.commit()
                book.close()
            except:
                print("엑셀오류")
                raise exceptions.ParseError("Excel calculation error")

    return HttpResponse(status=201)


@api_view(['GET', 'POST'])
@csrf_exempt
def cc_materialcost_if(request):
    if request.method == 'GET':  # Excel Template Download

        strsql1 = "SHOW TABLES LIKE 'cc_materialcost_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        rsColumns = None
        if rsTmp:
            strsql1 = "SHOW FULL COLUMNS FROM cc_materialcost_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

            idx = 1

            bookin = Workbook()
            sheet_in = bookin.active

            for i in rsColumns:
                sheet_in.cell(row=1, column=idx).value = i[1]
                sheet_in.cell(row=2, column=idx).value = i[8]
                idx += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="materialcost.xlsx"'

            bookin.save(response)
            bookin.close()

        return response

    elif request.method == 'POST':  # Excel Template Upload
        file = request.FILES['input_file']

        strsql1 = "SHOW TABLES LIKE 'cc_materialcost_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        if rsTmp:
            strsql1 = "SHOW COLUMNS FROM cc_materialcost_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

        max_col = len(rsColumns)
        book = openpyxl.load_workbook(file, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

        if max_row > 2:
            try:
                for j in range(3, max_row + 1):
                    lstTmp = []
                    strbottom = ""

                    for i in range(1, max_col + 1):
                        valTmp = sheet.cell(row=j, column=i).value

                        lstTmp.append(valTmp)

                        if valTmp == '':
                            strbottom += "default,"
                        elif valTmp == None:
                            strbottom += "default,"
                        else:
                            strbottom += "'" + str(valTmp) + "',"

                    strbottom = strbottom[:-1]

                    strSql = "INSERT INTO cc_materialcost_if VALUES (" + strbottom + ")"

                    cursor = dbCon.cursor()
                    cursor.execute(strSql)
                    rows = cursor.fetchone()
                    cursor.close()

                dbCon.commit()
                book.close()
            except:
                print("엑셀오류")
                raise exceptions.ParseError("Excel calculation error")

    return HttpResponse(status=201)


@api_view(['GET', 'POST'])
@csrf_exempt
def cc_itempermanucost_if(request):
    if request.method == 'GET':  # Excel Template Download

        strsql1 = "SHOW TABLES LIKE 'cc_itempermanucost_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        rsColumns = None
        if rsTmp:
            strsql1 = "SHOW FULL COLUMNS FROM cc_itempermanucost_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

            idx = 1

            bookin = Workbook()
            sheet_in = bookin.active

            for i in rsColumns:
                sheet_in.cell(row=1, column=idx).value = i[1]
                sheet_in.cell(row=2, column=idx).value = i[8]
                idx += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="itempermanucost.xlsx"'

            bookin.save(response)
            bookin.close()

        return response

    elif request.method == 'POST':  # Excel Template Upload
        file = request.FILES['input_file']

        strsql1 = "SHOW TABLES LIKE 'cc_itempermanucost_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        if rsTmp:
            strsql1 = "SHOW COLUMNS FROM cc_itempermanucost_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

        max_col = len(rsColumns)
        book = openpyxl.load_workbook(file, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

        if max_row > 2:
            try:
                for j in range(3, max_row + 1):
                    lstTmp = []
                    strbottom = ""

                    for i in range(1, max_col + 1):
                        valTmp = sheet.cell(row=j, column=i).value

                        lstTmp.append(valTmp)

                        if valTmp == '':
                            strbottom += "default,"
                        elif valTmp == None:
                            strbottom += "default,"
                        else:
                            strbottom += "'" + str(valTmp) + "',"

                    strbottom = strbottom[:-1]

                    strSql = "INSERT INTO cc_itempermanucost_if VALUES (" + strbottom + ")"

                    cursor = dbCon.cursor()
                    cursor.execute(strSql)
                    rows = cursor.fetchone()
                    cursor.close()

                dbCon.commit()
                book.close()
            except:
                print("엑셀오류")
                raise exceptions.ParseError("Excel calculation error")

    return HttpResponse(status=201)


@api_view(['GET', 'POST'])
@csrf_exempt
def cc_productcostpayment_if(request):
    if request.method == 'GET':  # Excel Template Download

        strsql1 = "SHOW TABLES LIKE 'cc_productcostpayment_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        rsColumns = None
        if rsTmp:
            strsql1 = "SHOW FULL COLUMNS FROM cc_productcostpayment_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

            idx = 1

            bookin = Workbook()
            sheet_in = bookin.active

            for i in rsColumns:
                sheet_in.cell(row=1, column=idx).value = i[1]
                sheet_in.cell(row=2, column=idx).value = i[8]
                idx += 1

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="productcostpayment.xlsx"'

            bookin.save(response)
            bookin.close()

        return response

    elif request.method == 'POST':  # Excel Template Upload
        file = request.FILES['input_file']

        strsql1 = "SHOW TABLES LIKE 'cc_productcostpayment_if'"

        cursor1 = dbCon.cursor()
        cursor1.execute(strsql1)
        rsTmp = cursor1.fetchone()
        cursor1.close()

        if rsTmp:
            strsql1 = "SHOW COLUMNS FROM cc_productcostpayment_if"

            cursor2 = dbCon.cursor()
            cursor2.execute(strsql1)
            rsColumns = cursor2.fetchall()
            cursor2.close()

        max_col = len(rsColumns)
        book = openpyxl.load_workbook(file, read_only=True)
        sheet = book.active
        max_row = sheet.max_row

        if max_row > 2:
            try:
                for j in range(3, max_row + 1):
                    lstTmp = []
                    strbottom = ""

                    for i in range(1, max_col + 1):
                        valTmp = sheet.cell(row=j, column=i).value

                        lstTmp.append(valTmp)

                        if valTmp == '':
                            strbottom += "default,"
                        elif valTmp == None:
                            strbottom += "default,"
                        else:
                            strbottom += "'" + str(valTmp) + "',"

                    strbottom = strbottom[:-1]

                    strSql = "INSERT INTO cc_productcostpayment_if VALUES (" + strbottom + ")"

                    cursor = dbCon.cursor()
                    cursor.execute(strSql)
                    rows = cursor.fetchone()
                    cursor.close()

                dbCon.commit()
                book.close()
            except:
                print("엑셀오류")
                raise exceptions.ParseError("Excel calculation error")

    return HttpResponse(status=201)


@api_view(['GET', 'POST'])
@csrf_exempt
def costbill_list(request):
    if request.method == 'GET':
        query_set = CcCostBill.objects.raw("SELECT * FROM cc_costbill")
        serializer = CcCostBillSerializer(query_set, many=True)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)


@api_view(['GET', 'PUT', 'DELETE'])
@csrf_exempt
def costbill_detail(request, pk):
    obj = CbCostCenter.objects.get(id=pk)

    if request.method == 'GET':  # 현재 화면에선 개별조회 미지원.
        serializer = CbCostCenterSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'PUT':
        data = JSONParser().parse(request)

        if CbCostCenter.objects.filter(cstctr_nm=data['cstctr_nm'], usage_fg='Y').exists():
            raise exceptions.ParseError("Duplicate Name")

        serializer = CbCostCenterSerializer(obj, data=data)

        if serializer.is_valid():
            serializer.save()
            return JsonResponse(serializer.data, status=200)
        return JsonResponse(serializer.errors, status=400)

    elif request.method == 'DELETE':
        obj.usage_fg = 'N'
        obj.save()
        return HttpResponse(status=204)


#####################ca_prediction##########################
@api_view(['GET', 'POST'])
@csrf_exempt
def ca_prediction(request):
    if request.method == 'GET':  # 가장 최근에 생성된 row 겍체를 반환.
        obj = CaPrediction.objects.last()
        serializer = CaPredictionSerializer(obj)
        return JsonResponse(serializer.data, safe=False)

    elif request.method == 'POST':
        data = JSONParser().parse(request)

        obj = CaPrediction.objects.filter(variableperc_cost=data['variableperc_cost'],
                                          fixedperc_cost=data['fixedperc_cost'],
                                          materialperc_cost=data['materialperc_cost'])
        if obj.exists():  # 변동비, 고정비, 재료비가 같은 row가 존재하면 해당 row 객체 반환.
            serializer = CaPredictionSerializer(obj)
            return JsonResponse(serializer.data, status=201)

        serializer = CaPredictionSerializer(data=data)  # 새로운 row 객체를 생성해서 변동비, 고정비, 재료비 속성에 입력 받은 값을 넣음.

        if serializer.is_valid():
            serializer.save()
            fData = CaPrediction.objects.filter(prediction4_cost=0, periodym4_cd=0, variableperc_cost=0, fixedperc_cost=0,
                                                 materialperc_cost=0, prediction4_max=0).first()

            rsData = CaPrediction.objects.last()
            rsData.prediction1_cost= fData.prediction1_cost
            rsData.periodym1_cd=fData.periodym1_cd
            rsData.prediction1_max = fData.prediction1_max
            rsData.prediction1_min = fData.prediction1_min
            rsData.save()

            a1=rsData.variableperc_cost
            a2=rsData.fixedperc_cost
            a3=rsData.materialperc_cost

            print("모델 동작 시키는 view 함수실행.")
            x1,x2,x3,x4=simulatorLoader(a1,a2,a3)
            print(x1,x2,x3,x4)
            rsData.periodym2_cd=x1[0]
            rsData.prediction2_cost= x2[0]
            rsData.prediction2_max = x3[0]
            rsData.prediction2_min = x4[0]

            rsData.periodym3_cd=x1[1]
            rsData.prediction3_cost= x2[1]
            rsData.prediction3_max = x3[1]
            rsData.prediction3_min = x4[1]

            rsData.periodym4_cd=x1[2]
            rsData.prediction4_cost= x2[2]
            rsData.prediction4_max = x3[2]
            rsData.prediction4_min = x4[2]
            rsData.save()
            return JsonResponse(serializer.data, status=201)

    return JsonResponse(serializer.errors, status=400)


####################train_data##################
@api_view(['POST'])
@csrf_exempt
def train_data(request):
    if request.method == 'POST':
        modelTrain()
        return HttpResponse(status=201)


####################predict_data#####################3
@api_view(['POST'])
@csrf_exempt
def predict_data(request):
    if request.method == 'POST':
        # print("예측")
        mainChartPredict()
        return HttpResponse(status=201)
